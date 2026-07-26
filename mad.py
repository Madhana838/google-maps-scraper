import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# Importing libraries
from selenium import webdriver
from selenium.webdriver.edge.service import Service as EdgeService   # Selenium 4: driver service object
from selenium.webdriver.edge.options import Options as EdgeOptions   # Edge-specific options class
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    StaleElementReferenceException,
)
import openpyxl
import time
import os
from datetime import datetime

# ---------------------------------------------------------------------------
# CONFIG (kept as variables instead of magic strings scattered in the script)
# ---------------------------------------------------------------------------
SEARCH_LOCATION = "Chennai, Tamil Nadu, India"
SEARCH_QUERY = "event organizers"
EXCEL_FILE = "event_organizers_chennai.xlsx"
SHEET_NAME = "Event Organizers"
MAX_SCROLLS = 30        # safety cap so the scroll loop can never run forever
SCROLL_PAUSE = 2        # seconds to wait after each scroll for new cards to load
MAX_RESULTS = 15         # only process this many businesses (set to None for all)

# ---------------------------------------------------------------------------
# DRIVER SETUP
# CHANGE: Firefox -> Microsoft Edge, using Selenium 4's Service/Options pattern
# instead of the old Selenium 3 style (webdriver.Firefox() with no options).
# ---------------------------------------------------------------------------
edge_options = EdgeOptions()
edge_options.add_argument("--start-maximized")
edge_options.add_argument("--disable-notifications")
edge_options.add_argument("--log-level=3")
# edge_options.add_argument("--headless=new")  # uncomment to run without a visible window

# Selenium 4 can auto-locate msedgedriver if it's on PATH and matches your
# installed Edge version. If you need a specific driver binary, uncomment:
# service = EdgeService(executable_path=r"C:\path\to\msedgedriver.exe")
# driver = webdriver.Edge(service=service, options=edge_options)
driver = webdriver.Edge(options=edge_options)
wait = WebDriverWait(driver, 15)


CONSENT_BUTTON_XPATH = (
    "//button[.//span[contains(text(),'Accept all')] "
    "or contains(@aria-label,'Accept all') "
    "or .//span[contains(text(),'I agree')] "
    "or contains(@aria-label,'I agree')]"
)


def dismiss_consent():
    """
    Google shows a consent screen before Maps loads, in one of two forms:
      1. An iframe embedded in the maps.google.com page (older/EU flow), or
      2. A full top-level redirect to consent.google.com (common in India and
         several other regions) with NO iframe at all.
    CHANGE: the old code only checked case 1. It's now checked first, and if
    no iframe/button is found, we also check the top-level page itself for a
    consent button before giving up. The old By.ID "introAgreeButton" locator
    is gone entirely from both flows, so a text/aria-label XPath is used.
    """
    # Case 1: iframe-based consent embedded in the maps page
    # CHANGE (bug fix): previously only "except TimeoutException" reset the
    # frame context. Maps pages can contain OTHER unrelated iframes (ads,
    # translate widgets, etc). If switching into one of those raised any
    # other exception, Selenium stayed stuck inside it forever, causing every
    # later find_element(By.ID, "searchboxinput") to time out even though the
    # page visually looked completely normal. Now every frame is wrapped in
    # try/except/finally so we ALWAYS return to the top-level document.
    try:
        iframes = driver.find_elements(By.TAG_NAME, "iframe")
        for frame in iframes:
            try:
                driver.switch_to.frame(frame)
                consent_button = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((By.XPATH, CONSENT_BUTTON_XPATH))
                )
                consent_button.click()
                time.sleep(2)
                return
            except Exception:
                pass
            finally:
                driver.switch_to.default_content()
    except Exception:
        driver.switch_to.default_content()

    # Case 2: full-page top-level redirect to consent.google.com (no iframe)
    if "consent.google.com" in driver.current_url or "consent" in driver.current_url.lower():
        try:
            consent_button = WebDriverWait(driver, 6).until(
                EC.element_to_be_clickable((By.XPATH, CONSENT_BUTTON_XPATH))
            )
            consent_button.click()
            time.sleep(2)
        except TimeoutException:
            print(f"Could not auto-dismiss consent page. Current URL: {driver.current_url}")


def scroll_and_collect_links():
    """
    Scrolls the Google Maps results panel until no new businesses load
    (or MAX_SCROLLS is hit), and returns the set of place links found.

    CHANGE: instead of clicking each search-result card in place (which
    constantly threw StaleElementReferenceException once Maps re-rendered
    the list), we collect every place URL first, then visit each URL
    directly. This is far more stable across Maps' current React-based DOM.
    """
    feed = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div[role="feed"]')))
    collected_links = set()
    previous_count = 0
    stagnant_rounds = 0

    for _ in range(MAX_SCROLLS):
        cards = feed.find_elements(By.CSS_SELECTOR, "a.hfpxzc")  # each card's anchor = place link
        for card in cards:
            href = card.get_attribute("href")
            if href:
                collected_links.add(href)

        driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", feed)
        time.sleep(SCROLL_PAUSE)

        # Stop once Google shows its "end of list" marker
        if feed.find_elements(By.XPATH, ".//span[contains(text(), \"You've reached the end of the list\")]"):
            break

        # Stop if two consecutive scrolls added no new links (list has stopped growing)
        if len(collected_links) == previous_count:
            stagnant_rounds += 1
            if stagnant_rounds >= 2:
                break
        else:
            stagnant_rounds = 0
        previous_count = len(collected_links)

    return list(collected_links)


def find_search_box(timeout=15):
    """
    Locate the Maps search input, trying a plain By.ID first and falling
    back to a shadow-DOM-piercing JavaScript search.

    CHANGE (root cause fix): Google has been migrating Maps UI pieces onto
    Web Components (custom elements like <gmp-map>) which encapsulate their
    markup inside a shadow root. A shadow root is NOT reachable by a normal
    By.ID/CSS selector from the top-level document — the element can be
    fully visible on screen while Selenium's plain locators still time out
    trying to find it, which matches exactly what was happening here. This
    function recursively pierces into any shadow roots on the page to find
    the real <input> element, and falls back to that if the plain ID lookup
    fails.
    """
    end_time = time.time() + timeout
    while time.time() < end_time:
        # 1) Fast path: plain top-level lookup (works if not inside shadow DOM)
        try:
            el = driver.find_element(By.ID, "searchboxinput")
            if el.is_displayed():
                return el
        except NoSuchElementException:
            pass

        # 2) Fallback: recursively search through shadow roots via JS
        el = driver.execute_script("""
            function findInput(root) {
                let el = root.querySelector(
                    'input#searchboxinput, input[aria-label="Search Google Maps"], input[name="q"]'
                );
                if (el) return el;
                const nodes = root.querySelectorAll('*');
                for (const node of nodes) {
                    if (node.shadowRoot) {
                        const found = findInput(node.shadowRoot);
                        if (found) return found;
                    }
                }
                return null;
            }
            return findInput(document);
        """)
        if el is not None:
            return el

        time.sleep(0.5)

    raise TimeoutException("Could not locate the Maps search box (plain DOM or shadow DOM).")


def extract_text(by, selector, timeout=6):
    """Small helper: wait for an element and return its stripped text, or a fallback string."""
    try:
        el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, selector)))
        text = el.text.strip()
        return text if text else "Not available"
    except (TimeoutException, NoSuchElementException):
        return "Not available"


# ---------------------------------------------------------------------------
# MAIN SCRIPT
# ---------------------------------------------------------------------------
business_data = []

try:
    driver.get("https://www.google.com/maps")
    time.sleep(3)
    dismiss_consent()

    # Perform a single combined search instead of "location" then "clear + query".
    # CHANGE: searching location first and then clearing/re-typing the query now
    # often lands on a location place-page where the search box behaves
    # differently, breaking the old two-step flow. One combined query is stable.
    try:
        # Defensive reset: guarantees we're never accidentally still inside
        # an iframe context left over from dismiss_consent() or anything else.
        driver.switch_to.default_content()

        # CHANGE: replaced the plain By.ID wait with find_search_box(), which
        # also pierces shadow DOM if Google has wrapped the input in a
        # Web Component (see function docstring for why this was needed).
        search_box = find_search_box(timeout=15)
        search_box.clear()
        search_box.send_keys(f"{SEARCH_QUERY} in {SEARCH_LOCATION}")
        search_box.send_keys(Keys.ENTER)

        # Wait for the actual results feed instead of an unconditional time.sleep
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div[role="feed"]')))
    except (TimeoutException, NoSuchElementException) as e:
        # CHANGE: on failure, dump the page title/URL and a screenshot instead
        # of just re-raising Selenium's generic (and unhelpful) stacktrace.
        # This tells you immediately whether Maps loaded, got stuck on a
        # consent/redirect page, or something else entirely.
        print("Error while interacting with search box.")
        print(f"Current URL: {driver.current_url}")
        print(f"Page title: {driver.title}")
        try:
            driver.save_screenshot("debug_search_box_failure.png")
            print("Saved screenshot to debug_search_box_failure.png for inspection.")
        except Exception:
            pass
        raise

    # Collect all business links by scrolling the results feed
    links = scroll_and_collect_links()
    print(f"Number of results found: {len(links)}")

    # CHANGE: cap how many businesses actually get processed/saved, since the
    # feed can return far more results than needed (e.g. 80 found, only want 15).
    if MAX_RESULTS is not None:
        links = links[:MAX_RESULTS]
        print(f"Processing only the first {len(links)} result(s) (MAX_RESULTS={MAX_RESULTS}).")

    # Visit each business's own Maps URL to read its detail panel
    for index, link in enumerate(links, start=1):
        try:
            print(f"Processing result {index}/{len(links)}...")
            driver.get(link)

            # Business name: current Maps renders it as an <h1> in the detail panel
            name = extract_text(By.CSS_SELECTOR, "h1.DUwDvf", timeout=10)

            # Address: CHANGE: old code matched any div with class "Io6YTe" that
            # contained a comma, which is fragile since that class is reused for
            # many unrelated fields. Now scoped to the address button via the
            # stable data-item-id="address" attribute Google actually uses.
            address = extract_text(By.CSS_SELECTOR, 'button[data-item-id="address"] div.Io6YTe')

            # Phone: CHANGE: old code matched any "Io6YTe" div containing "+",
            # which breaks for numbers without a "+" prefix and can match the
            # wrong field. Now scoped via data-item-id starting with "phone:tel:".
            phone = extract_text(By.CSS_SELECTOR, 'button[data-item-id^="phone:tel:"] div.Io6YTe')

            print(f"Extracted: {name} | {address} | {phone}")
            business_data.append([SEARCH_LOCATION, name, address, phone, link])

        except (TimeoutException, NoSuchElementException, StaleElementReferenceException) as e:
            print(f"Error processing result {index}: {e}")
            business_data.append([SEARCH_LOCATION, "Error", "Error", "Error", link])

finally:
    # Always close the browser, even if something above raised an exception
    driver.quit()

# ---------------------------------------------------------------------------
# EXCEL OUTPUT (unchanged logic, kept from the original script)
# ---------------------------------------------------------------------------
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    wb.save(EXCEL_FILE)

wb = openpyxl.load_workbook(EXCEL_FILE)
if SHEET_NAME not in wb.sheetnames:
    sheet = wb.create_sheet(SHEET_NAME)
else:
    sheet = wb[SHEET_NAME]

if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(1, 1).value is None:
    sheet.append(["Location", "Business Name", "Address", "Phone Number", "Link"])

for row in business_data:
    sheet.append(row)

try:
    wb.save(EXCEL_FILE)
    print(f"Data successfully saved to {EXCEL_FILE}")
except PermissionError:
    print(f"PermissionError: Unable to save to {EXCEL_FILE}. Ensure the file is not open or locked.")
    backup_file = "event_organizers_latur_backup.xlsx"
    try:
        wb.save(backup_file)
        print(f"Data successfully saved to backup file: {backup_file}")
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file_dynamic = f"event_organizers_latur_backup_{timestamp}.xlsx"
        wb.save(backup_file_dynamic)
        print(f"Data successfully saved to a dynamic backup file: {backup_file_dynamic}")

